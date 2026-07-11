#!/usr/bin/env python3
"""Extract betting pool picks from the Google Sheets xlsx export into data.js.

Usage: python3 extract_picks.py "/path/to/Jalkapallon mm-kisaveikkaus 2026.xlsx"
"""
import json
import re
import sys
import unicodedata
from datetime import datetime, timedelta, timezone

import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "/Users/kailanto/Downloads/Jalkapallon mm-kisaveikkaus 2026.xlsx"

# Finnish sheet name -> ESPN displayName
TEAM_MAP = {
    "meksiko": "Mexico", "etelä-afrikka": "South Africa", "etelä-korea": "South Korea",
    "tšekki": "Czechia", "tsekki": "Czechia", "kanada": "Canada",
    "bosnia ja hertsegovina": "Bosnia-Herzegovina", "usa": "United States",
    "paraguay": "Paraguay", "brasilia": "Brazil", "marokko": "Morocco",
    "australia": "Australia", "turkki": "Türkiye", "haiti": "Haiti",
    "skotlanti": "Scotland", "qatar": "Qatar", "sveitsi": "Switzerland",
    "saksa": "Germany", "curaçao": "Curaçao", "curacao": "Curaçao",
    "norsunluurannikko": "Ivory Coast", "hollanti": "Netherlands",
    "alankomaat": "Netherlands", "japani": "Japan", "ruotsi": "Sweden",
    "tunisia": "Tunisia", "espanja": "Spain", "kap verde": "Cape Verde",
    "belgia": "Belgium", "egypti": "Egypt", "saudi-arabia": "Saudi Arabia",
    "iran": "Iran", "uusi-seelanti": "New Zealand", "ranska": "France",
    "senegal": "Senegal", "irak": "Iraq", "norja": "Norway",
    "argentiina": "Argentina", "argentina": "Argentina", "jordania": "Jordan",
    "itävalta": "Austria", "algeria": "Algeria", "portugali": "Portugal",
    "portugal": "Portugal", "uzbekistan": "Uzbekistan", "englanti": "England",
    "kroatia": "Croatia", "ghana": "Ghana", "panama": "Panama",
    "kolumbia": "Colombia", "kongon dr": "Congo DR", "ecuador": "Ecuador",
    "uruguay": "Uruguay",
    # common typos in pick cells
    "braslia": "Brazil", "espanija": "Spain", "engalnti": "England",
    "norsunluurannikk": "Ivory Coast", "japan": "Japan", "meksico": "Mexico",
}

def norm(s):
    s = unicodedata.normalize("NFC", str(s)).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def map_team(s, context=""):
    n = norm(s)
    if n in TEAM_MAP:
        return TEAM_MAP[n]
    # prefix match for truncated entries
    for k, v in TEAM_MAP.items():
        if n.startswith(k) or k.startswith(n):
            return v
    print(f"WARN: unmapped team {s!r} ({context})", file=sys.stderr)
    return None

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Veikkausrivit"]

# participants: name in row 2 at odd columns 3..41
participants = []
pick_cols = []
for col in range(3, 43, 2):
    name = ws.cell(row=2, column=col).value
    if name:
        participants.append(str(name).strip())
        pick_cols.append(col)

# Finnish summer time (EEST) = UTC+3
EEST = timezone(timedelta(hours=3))

matches = []
specials = {"top8": {}, "top4": {}, "semiScorer": {}, "topScorer": {},
            "bronze": {}, "finalLoser": {}, "champion": {}}
cur_date = None
date_re = re.compile(r"^[A-Za-zÄÖÅäöå]{2}\s+(\d{1,2})\.(\d{1,2})\.")
match_re = re.compile(r"^(\d{1,2})\.(\d{2})\s+(.+?)\s+–\s+(.+?)\s*\(([A-L])\)\s*$")
score_re = re.compile(r"^\s*(\d+)\s*[-–]\s*(\d+)\s*$")

def parse_score(v):
    if v is None:
        return None
    m = score_re.match(str(v))
    return [int(m.group(1)), int(m.group(2))] if m else None

for r in range(4, 94):
    a = ws.cell(row=r, column=1).value
    if a is None:
        continue
    a = str(a).strip()
    dm = date_re.match(a)
    if dm:
        cur_date = (int(dm.group(1)), int(dm.group(2)))
        continue
    mm = match_re.match(a)
    if not mm:
        if a:
            print(f"WARN: row {r} unparsed: {a!r}", file=sys.stderr)
        continue
    hh, mins, home_fi, away_fi, group = mm.groups()
    day, month = cur_date
    ko = datetime(2026, month, day, int(hh), int(mins), tzinfo=EEST)
    home = map_team(home_fi, f"row {r}")
    away = map_team(away_fi, f"row {r}")
    result = parse_score(ws.cell(row=r, column=2).value)
    picks = []
    for col in pick_cols:
        picks.append(parse_score(ws.cell(row=r, column=col).value))
    matches.append({
        "row": r, "group": group,
        "homeFi": home_fi.strip(), "awayFi": away_fi.strip(),
        "home": home, "away": away,
        "kickoff": ko.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%MZ"),
        "sheetResult": result,
        "picks": picks,
    })

# specials
def team_picks(rows):
    out = {p: [] for p in participants}
    for r in rows:
        for p, col in zip(participants, pick_cols):
            v = ws.cell(row=r, column=col).value
            if v is None or not str(v).strip():
                continue
            t = map_team(v, f"special row {r} {p}")
            out[p].append(t or str(v).strip())
    return out

# clear misspellings of scorer names -> corrected (matched on the surname token)
SCORER_TYPO = {"mbabbe": "Mbappé", "mbappe": "Mbappé"}

def fix_scorer(v):
    v = str(v or "").strip()
    if not v:
        return None
    toks = v.split()
    fixed = [SCORER_TYPO.get(t.lower(), t) for t in toks]
    return " ".join(fixed)

def text_picks(r):
    return {p: fix_scorer(ws.cell(row=r, column=col).value)
            for p, col in zip(participants, pick_cols)}

def single_team_picks(r):
    out = {}
    for p, col in zip(participants, pick_cols):
        v = ws.cell(row=r, column=col).value
        out[p] = map_team(v, f"special row {r} {p}") if v else None
    return out

specials["top8"] = team_picks(range(95, 103))
specials["top4"] = team_picks(range(104, 108))
specials["semiScorer"] = text_picks(109)
specials["topScorer"] = text_picks(110)
specials["bronze"] = single_team_picks(111)
specials["finalLoser"] = single_team_picks(112)
specials["champion"] = single_team_picks(113)

data = {
    "title": "MM-Mundo 2026",
    "prizes": "Voitonjako: 1. 100€, 2. 60€, 3. 40€.",
    "participants": participants,
    "matches": matches,
    "specials": specials,
}

with open("data.js", "w") as f:
    f.write("// Generated by extract_picks.py from the betting pool xlsx — do not edit by hand\n")
    f.write("const POOL_DATA = ")
    json.dump(data, f, ensure_ascii=False, indent=1)
    f.write(";\n")

print(f"OK: {len(participants)} participants, {len(matches)} matches -> data.js")
